from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .config import norm_text


def read_table(path: str | Path, expected_fields: list[str] | None = None) -> pd.DataFrame:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        df = _read_excel(path, expected_fields or [])
    elif suffix == ".csv":
        df = _read_csv(path)
    else:
        raise ValueError(f"不支持的文件格式：{path.suffix}")
    df = df.dropna(how="all")
    df.columns = _clean_columns(df.columns)
    df = df.loc[:, [not str(col).startswith("Unnamed:") for col in df.columns]]
    return df


def _read_excel(path: Path, expected_fields: list[str]) -> pd.DataFrame:
    if path.suffix.lower() == ".xls":
        return pd.read_excel(path, dtype=str)

    raw = pd.read_excel(path, header=None, dtype=str)
    raw = _drop_empty_edges(raw)
    if raw.empty:
        return raw

    expected = {norm_text(item) for item in expected_fields if norm_text(item)}
    if len(raw) < 2 or not expected:
        df = pd.read_excel(path, dtype=str)
        return df.dropna(how="all")

    row_1 = [norm_text(value) for value in raw.iloc[0].tolist()]
    row_2 = [norm_text(value) for value in raw.iloc[1].tolist()]
    row_1_matches = sum(1 for value in row_1 if value in expected)
    row_2_matches = sum(1 for value in row_2 if value in expected)
    merged_header = _has_merged_header(path)

    if row_1_matches + row_2_matches == 0:
        df = pd.read_excel(path, dtype=str)
        return df.dropna(how="all")

    header_rows = 2 if row_2_matches > 0 or merged_header else 1
    names = []
    for col in range(raw.shape[1]):
        values = [norm_text(raw.iat[row, col]) for row in range(min(header_rows, len(raw)))]
        exact = [value for value in values if value in expected]
        if exact:
            name = exact[-1]
        else:
            parts = []
            for value in values:
                if value and value not in parts:
                    parts.append(value)
            name = " ".join(parts)
        names.append(name)

    df = raw.iloc[header_rows:].copy()
    df.columns = names
    return df.dropna(how="all")


def _drop_empty_edges(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(how="all")
    df = df.dropna(axis=1, how="all")
    return df.reset_index(drop=True)


def _has_merged_header(path: Path) -> bool:
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
    except Exception:
        return False
    ws = wb.active
    for cell_range in ws.merged_cells.ranges:
        if cell_range.min_row <= 2 and cell_range.max_row >= 2:
            return True
    return False


def _read_csv(path: Path) -> pd.DataFrame:
    last_error: Exception | None = None
    for encoding in ["utf-8-sig", "gb18030", "gbk", "utf-8"]:
        try:
            return pd.read_csv(path, dtype=str, encoding=encoding)
        except Exception as exc:
            last_error = exc
    raise ValueError(f"CSV 读取失败：{path}") from last_error


def _clean_columns(columns: object) -> list[str]:
    seen: dict[str, int] = {}
    cleaned = []
    for raw in columns:
        name = norm_text(raw)
        if not name:
            name = "未命名字段"
        count = seen.get(name, 0)
        seen[name] = count + 1
        cleaned.append(name if count == 0 else f"{name}_{count + 1}")
    return cleaned
