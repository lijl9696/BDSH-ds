from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import PLATFORMS
from .paths import app_root
from .processor import ProcessResult


ROOT = app_root()
MEITUAN_TEMPLATE = ROOT / "配置表" / "美团输出报表模板.xlsx"
DOUYIN_TEMPLATE = ROOT / "配置表" / "抖音输出报表模板.xlsx"
DEFAULT_TEMPLATE = MEITUAN_TEMPLATE
MEITUAN_LOGO = ROOT / "assets" / "excel" / "meituan.png"
DOUYIN_LOGO = ROOT / "assets" / "excel" / "dy.png"
MEITUAN_YELLOW = "FFD100"
DOUYIN_BLACK = "000000"
DEFAULT_FONT_NAME = "Microsoft YaHei"
TRAFFIC_FIELDS = ["曝光人数(人)", "访问人数(人)", "下单人数(人)", "核销单量", "新客核销（人）", "新增评价", "新增好评"]
BUSINESS_LEVELS = ["金牌", "银牌", "铜牌", "无等级"]
SUMMARY_SHEETS = {"美团流量汇总", "美团经营汇总", "抖音汇总", "合并汇总"}
DIMENSION_FIELDS = {
    "平台",
    "统计日期",
    "省份",
    "城市",
    "城市负责人",
    "门店负责人",
    "大区",
    "大区负责人",
    "门店名",
    "门店名称",
    "门店ID",
    "分类字段",
    "分类值",
    "牌级",
    "门店牌级",
    "牌级别",
    "处理状态",
    "异常原因",
}
GROUP_RULES = {"分组", "维度", "纬度", "分类", "按此分组", "group", "groupby"}


FIELD_ALIASES = {
    "门店名称": "门店名",
    "门店": "门店名",
    "门店牌级": "牌级",
    "牌级别": "牌级",
    "曝光人数": "曝光人数(人)",
    "访问人数": "访问人数(人)",
    "下单人数": "下单人数(人)",
    "核销新客人数": "新客核销（人）",
    "新客核销人数": "新客核销（人）",
    "新增评价数": "新增评价",
    "评价数": "新增评价",
    "新增好评数": "新增好评",
    "好评数": "新增好评",
    "新增差评数": "新增差评",
    "差评数": "新增差评",
    "经营评分均值": "经营评分",
    "经营评分平均值": "经营评分",
    "经营评分最高值": "经营评分",
    "经营评分最低值": "经营评分",
    "订单数": "核销单量",
}


METHOD_WORDS = {
    "求和",
    "合计",
    "汇总",
    "sum",
    "total",
    "取整",
    "整数",
    "round",
    "平均",
    "平均值",
    "均值",
    "avg",
    "mean",
    "最高",
    "最大",
    "max",
    "最低",
    "最小",
    "min",
    "去重计数",
    "不重复计数",
    "唯一计数",
    "计数",
    "记录数",
    "行数",
    "原值",
    "保留",
    "first",
}


@dataclass(frozen=True)
class OutputSpec:
    headers: list[str]
    rules: list[str]
    header_row: int
    rule_row: int | None


@dataclass(frozen=True)
class RuleSpec:
    source: str
    method: str
    conditions: list[tuple[str, str]]
    is_group: bool = False


def write_excel_report(result: ProcessResult, output_path: str | Path, template_path: str | Path | None = None) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    template = _resolve_template(template_path, DEFAULT_TEMPLATE)
    if template:
        wb = load_workbook(template)
        if not wb.sheetnames:
            wb = Workbook()
    else:
        wb = Workbook()
        wb.active.title = "合并明细"

    template_sheet_names = [name for name in wb.sheetnames if name != "处理说明"]
    specs = {name: _read_output_spec(wb, name) for name in template_sheet_names}
    platform_key = _single_platform(result)
    for name in template_sheet_names:
        df = _build_output_sheet(name, _source_for_sheet(name, result), result, specs.get(name))
        if _uses_platform_header(name, platform_key):
            _write_platform_header_sheet(wb, name, df, result, platform_key or _platform_from_sheet(name))
        else:
            _write_sheet(wb, name, df)

    if "处理说明" not in wb.sheetnames:
        ws = wb.create_sheet("处理说明")
    else:
        ws = wb["处理说明"]
        _clear_sheet(ws)
    ws.append(["统计开始日期", result.period_start])
    ws.append(["统计结束日期", result.period_end])
    ws.append(["已导入平台", "、".join(PLATFORMS.get(key, key) for key in result.details_by_platform)])
    ws.append(["汇总指标", "、".join(_metric_label(field, result.aggregate_map.get(field, "sum")) for field in result.summary_fields)])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    _apply_default_font(wb)
    wb.save(output_path)
    return output_path


def write_platform_excel_reports(
    result: ProcessResult,
    output_dir: str | Path,
    timestamp: str | None = None,
    meituan_template: str | Path | None = None,
    douyin_template: str | Path | None = None,
) -> list[Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = timestamp or pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    paths: list[Path] = []
    if "meituan" in result.details_by_platform:
        template = _resolve_template(meituan_template, MEITUAN_TEMPLATE)
        paths.append(write_excel_report(_result_for_platform(result, "meituan"), output_dir / f"美团报表输出_{timestamp}.xlsx", template))
    if "douyin" in result.details_by_platform:
        douyin_only = _result_for_platform(result, "douyin")
        template = _resolve_template(douyin_template, DOUYIN_TEMPLATE)
        if template:
            paths.append(write_excel_report(douyin_only, output_dir / f"抖音报表输出_{timestamp}.xlsx", template))
        else:
            paths.append(write_excel_report(douyin_only, output_dir / f"抖音报表输出_{timestamp}.xlsx", DOUYIN_TEMPLATE))
    return paths


def build_meituan_traffic_summary(result: ProcessResult) -> pd.DataFrame:
    df = result.details_by_platform.get("meituan", pd.DataFrame()).copy()
    columns = ["平台", "大区", "门店负责人", "门店名称", *TRAFFIC_FIELDS]
    if df.empty:
        return pd.DataFrame(columns=columns)
    for col in ["平台", "大区", "门店负责人", "门店名"]:
        if col not in df.columns:
            df[col] = ""
    for field in TRAFFIC_FIELDS:
        df[field] = _number_column(df, field)
    summary = (
        df.groupby(["平台", "大区", "门店负责人", "门店名"], dropna=False)[TRAFFIC_FIELDS]
        .sum()
        .reset_index()
        .rename(columns={"门店名": "门店名称"})
    )
    return _order_existing(summary, columns)


def build_meituan_business_summary(result: ProcessResult) -> pd.DataFrame:
    df = result.details_by_platform.get("meituan", pd.DataFrame()).copy()
    columns = ["平台", "大区", "大区负责人", "分类字段", "分类值", "门店数", "新增评价", "新增好评", "经营评分平均值", "经营评分最高值"]
    if df.empty:
        return pd.DataFrame(columns=columns)
    for col in ["平台", "大区", "大区负责人", "城市负责人", "门店负责人", "门店名", "牌级别"]:
        if col not in df.columns:
            df[col] = ""
    df["大区负责人"] = _first_non_empty(df, ["大区负责人", "城市负责人", "门店负责人"])
    df["分类字段"] = "牌级别"
    df["分类值"] = df["牌级别"].map(_business_level)
    df["新增评价"] = _number_column(df, "新增评价")
    df["新增好评"] = _number_column(df, "新增好评")
    df["经营评分"] = _number_column(df, "经营评分")
    rows: list[dict[str, object]] = []
    grouped = df.groupby(["平台", "大区", "大区负责人", "分类字段", "分类值"], dropna=False)
    for keys, group in grouped:
        platform, area, owner, category_field, category_value = keys
        scores = group["经营评分"][group["经营评分"] > 0]
        rows.append(
            {
                "平台": platform,
                "大区": area,
                "大区负责人": owner,
                "分类字段": category_field,
                "分类值": category_value,
                "门店数": group["门店名"].map(_text_value).replace("", pd.NA).nunique(),
                "新增评价": group["新增评价"].sum(),
                "新增好评": group["新增好评"].sum(),
                "经营评分平均值": scores.mean() if not scores.empty else 0,
                "经营评分最高值": scores.max() if not scores.empty else 0,
            }
        )
    summary = pd.DataFrame(rows, columns=columns)
    if summary.empty:
        return summary
    summary["_level_order"] = summary["分类值"].map({name: idx for idx, name in enumerate(BUSINESS_LEVELS)})
    summary = summary.sort_values(["大区", "大区负责人", "_level_order"]).drop(columns=["_level_order"])
    return summary


def build_douyin_summary(result: ProcessResult) -> pd.DataFrame:
    df = result.details_by_platform.get("douyin", pd.DataFrame()).copy()
    if df.empty:
        return pd.DataFrame(columns=["平台", "大区", "门店负责人", "记录数"])
    metrics = [field for field, method in result.aggregate_map.items() if method != "none" and field in df.columns]
    for metric in metrics:
        df[metric] = _number_column(df, metric)
    for col in ["平台", "大区", "门店负责人"]:
        if col not in df.columns:
            df[col] = ""
    if metrics:
        summary = df.groupby(["平台", "大区", "门店负责人"], dropna=False)[metrics].sum().reset_index()
    else:
        summary = df.groupby(["平台", "大区", "门店负责人"], dropna=False).size().reset_index(name="记录数")
    counts = df.groupby(["平台", "大区", "门店负责人"], dropna=False).size().reset_index(name="记录数")
    if "记录数" in summary.columns:
        summary = summary.drop(columns=["记录数"])
    return summary.merge(counts, on=["平台", "大区", "门店负责人"], how="left")


def _build_output_sheet(name: str, source: pd.DataFrame, result: ProcessResult, spec: OutputSpec | None) -> pd.DataFrame:
    if not spec:
        if name == "美团流量汇总":
            return build_meituan_traffic_summary(result)
        if name == "美团经营汇总":
            return build_meituan_business_summary(result)
        if name == "抖音汇总":
            return build_douyin_summary(result)
        if name == "合并汇总":
            return result.combined_summary
        return source
    return build_template_sheet(name, source, spec)


def build_template_sheet(name: str, source: pd.DataFrame, spec: OutputSpec) -> pd.DataFrame:
    data = _prepare_template_source(name, source)
    headers = spec.headers
    if data.empty:
        return pd.DataFrame(columns=headers)

    group_specs = _template_group_specs(data, spec)
    if group_specs:
        return _build_grouped_template(data, headers, spec.rules, group_specs)
    return _build_row_template(data, headers, spec.rules)


def build_template_summary(name: str, source: pd.DataFrame, spec: OutputSpec) -> pd.DataFrame:
    return build_template_sheet(name, source, spec)


def _build_grouped_template(data: pd.DataFrame, headers: list[str], rules: list[str], group_specs: list[tuple[str, RuleSpec]]) -> pd.DataFrame:
    for _, rule_spec in group_specs:
        data = _filter_data(data, rule_spec.conditions)
    if data.empty:
        return pd.DataFrame(columns=headers)

    group_fields: list[str] = []
    for header, rule_spec in group_specs:
        source = _source_field(data, rule_spec.source or header)
        if source not in data.columns:
            continue
        if header != source:
            data[header] = data[source]
        if header not in group_fields:
            group_fields.append(header)
    if not group_fields:
        return _project_template_columns(data, headers)

    grouped = data.groupby(group_fields, dropna=False, sort=False)
    output = grouped.size().reset_index(name="_记录数")
    row_count = len(output)

    base_keys = output[group_fields].copy()
    for header, rule in zip(headers, rules):
        if header in group_fields:
            continue
        if _is_formula_rule(rule):
            output[header] = _evaluate_formula(rule, output, grouped, group_fields, data, row_count)
        else:
            output[header] = _aggregate_template_column(grouped, group_fields, base_keys, data, header, _parse_rule(header, rule))

    output = output.drop(columns=["_记录数"], errors="ignore")
    return _project_template_columns(output, headers)


def _build_row_template(data: pd.DataFrame, headers: list[str], rules: list[str]) -> pd.DataFrame:
    output = pd.DataFrame(index=data.index)
    for header, rule in zip(headers, rules):
        rule_spec = _parse_rule(header, rule)
        if _is_formula_rule(rule):
            output[header] = _evaluate_row_formula(rule, output, data)
            continue
        filtered = _filter_data(data, rule_spec.conditions)
        source = _source_field(filtered, rule_spec.source or header)
        if source not in filtered.columns:
            output[header] = ""
            continue
        values = filtered[source].reindex(data.index)
        if rule_spec.method == "round":
            output[header] = pd.to_numeric(values, errors="coerce").fillna(0).round(0).astype("int64")
        else:
            output[header] = values
    return output.reset_index(drop=True)


def _read_output_spec(wb, name: str) -> OutputSpec | None:
    if name not in wb.sheetnames:
        return None
    ws = wb[name]
    preferred = _template_row(ws, 3)
    legacy = _template_row(ws, 4)
    if _usable_header_row(preferred):
        headers = preferred
        rules = _template_row(ws, 4, len(headers))
        rules = rules if not _looks_like_reference_row(rules) else []
        return OutputSpec(headers=headers, rules=_pad_rules(rules, len(headers)), header_row=3, rule_row=4)
    if _usable_header_row(legacy):
        headers = legacy
        candidate_rules = _template_row(ws, 5, len(headers))
        rules = candidate_rules if _looks_like_rule_row(candidate_rules) else []
        return OutputSpec(headers=headers, rules=_pad_rules(rules, len(headers)), header_row=4, rule_row=5 if rules else None)
    return None


def _template_row(ws, row_idx: int, width: int | None = None) -> list[str]:
    max_col = width or ws.max_column
    values = [_text_value(ws.cell(row_idx, col).value) for col in range(1, max_col + 1)]
    if width:
        return values
    last = 0
    for idx, value in enumerate(values, 1):
        if value:
            last = idx
    return values[:last]


def _usable_header_row(values: list[str]) -> bool:
    non_empty = [value for value in values if value]
    if len(non_empty) < 2:
        return False
    if _looks_like_reference_row(non_empty):
        return False
    first = non_empty[0]
    if "参考配置示例" in first or "不用" in first:
        return False
    return True


def _looks_like_reference_row(values: list[str]) -> bool:
    compact = [value for value in values if value]
    if not compact:
        return False
    if compact[0] in {"配置项", "用途", "参考说明"}:
        return True
    if "参考说明" in compact[:3]:
        return True
    return any("参考配置示例" in value for value in compact[:2])


def _looks_like_rule_row(values: list[str]) -> bool:
    if not values:
        return False
    matched = 0
    for value in values:
        rule = _norm_rule(value)
        if not rule:
            continue
        if rule in GROUP_RULES or _is_formula_rule(rule) or _rule_method(rule) != "":
            matched += 1
    return matched > 0


def _pad_rules(rules: list[str], length: int) -> list[str]:
    normalized = [_norm_rule(rule) for rule in rules[:length]]
    return normalized + [""] * (length - len(normalized))


def _prepare_template_source(name: str, source: pd.DataFrame) -> pd.DataFrame:
    data = source.copy()
    if "门店名称" not in data.columns and "门店名" in data.columns:
        data["门店名称"] = data["门店名"]
    if "门店名" not in data.columns and "门店名称" in data.columns:
        data["门店名"] = data["门店名称"]
    for col in ["平台", "大区", "大区负责人", "城市负责人", "门店负责人", "门店名", "牌级别", "牌级"]:
        if col not in data.columns:
            data[col] = ""
    data["大区负责人"] = _first_non_empty(data, ["大区负责人", "城市负责人", "门店负责人"])
    data["牌级"] = _first_non_empty(data, ["牌级", "牌级别"]).map(_business_level)
    data["牌级别"] = data["牌级"]
    data["门店牌级"] = data["牌级"]
    data["分类字段"] = "牌级"
    data["分类值"] = data["牌级"]
    if "新增差评" not in data.columns and {"新增评价", "新增好评"} <= set(data.columns):
        data["新增差评"] = _number_column(data, "新增评价") - _number_column(data, "新增好评")
    return data


def _template_group_specs(data: pd.DataFrame, spec: OutputSpec) -> list[tuple[str, RuleSpec]]:
    group_specs = []
    for header, rule in zip(spec.headers, spec.rules):
        rule_spec = _parse_rule(header, rule)
        if not rule_spec.is_group:
            continue
        source = _source_field(data, rule_spec.source or header)
        if source in data.columns or header in data.columns:
            group_specs.append((header, rule_spec))
    return group_specs


def _aggregate_template_column(grouped, group_fields: list[str], base_keys: pd.DataFrame, data: pd.DataFrame, header: str, rule_spec: RuleSpec) -> pd.Series:
    method = rule_spec.method
    source = _source_field(data, rule_spec.source or header)
    data = _filter_data(data, rule_spec.conditions)
    grouped = data.groupby(group_fields, dropna=False, sort=False) if not data.empty else None
    if method == "record_count":
        return _align_grouped_series(grouped.size() if grouped is not None else pd.Series(dtype="int64"), group_fields, base_keys)
    if method == "nunique" or header == "门店数":
        count_field = source if source in data.columns else "门店名"
        if grouped is not None and count_field in data.columns:
            series = grouped[count_field].agg(lambda values: values.map(_text_value).replace("", pd.NA).dropna().nunique())
            return _align_grouped_series(series, group_fields, base_keys)
        return pd.Series([0] * len(base_keys))
    if grouped is None or source not in data.columns:
        return pd.Series([0] * len(base_keys))
    if method == "first":
        return _align_grouped_series(grouped[source].first(), group_fields, base_keys)
    if method == "mean":
        values = pd.to_numeric(data[source], errors="coerce").fillna(0)
        temp = data.assign(**{source: values})
        if "评分" in source:
            temp[source] = temp[source].where(temp[source] > 0)
        series = temp.groupby(group_fields, dropna=False, sort=False)[source].mean().fillna(0)
        return _align_grouped_series(series, group_fields, base_keys)
    if method == "max":
        temp = data.assign(**{source: pd.to_numeric(data[source], errors="coerce").fillna(0)})
        return _align_grouped_series(temp.groupby(group_fields, dropna=False, sort=False)[source].max(), group_fields, base_keys)
    if method == "min":
        temp = data.assign(**{source: pd.to_numeric(data[source], errors="coerce").fillna(0)})
        return _align_grouped_series(temp.groupby(group_fields, dropna=False, sort=False)[source].min(), group_fields, base_keys)

    temp = data.assign(**{source: pd.to_numeric(data[source], errors="coerce").fillna(0)})
    summed = _align_grouped_series(temp.groupby(group_fields, dropna=False, sort=False)[source].sum(), group_fields, base_keys)
    if method == "round":
        return summed.round(0).astype("int64")
    if method == "":
        if not _looks_numeric(data[source]):
            return _align_grouped_series(grouped[source].first(), group_fields, base_keys)
    return summed


def _rule_method(rule: str) -> str:
    text = _norm_rule(rule)
    if not text:
        return ""
    if any(word in text.split() for word in GROUP_RULES):
        return "first"
    if any(word in text for word in ["记录数", "行数"]):
        return "record_count"
    if any(word in text for word in ["去重计数", "不重复计数", "唯一计数", "门店数"]):
        return "nunique"
    if any(word in text for word in ["平均", "均值", "avg", "mean"]):
        return "mean"
    if any(word in text for word in ["最大", "最高", "max"]):
        return "max"
    if any(word in text for word in ["最小", "最低", "min"]):
        return "min"
    if any(word in text for word in ["取整", "整数", "round"]):
        return "round"
    if any(word in text for word in ["求和", "合计", "汇总", "sum", "total"]):
        return "sum"
    if any(word in text for word in ["原值", "保留", "first"]):
        return "first"
    return ""


def _parse_rule(header: str, rule: str) -> RuleSpec:
    text = _normalize_formula_text(rule)
    if not text:
        return RuleSpec(source=header, method="", conditions=[])
    tokens = text.split()
    conditions: list[tuple[str, str]] = []
    source_parts: list[str] = []
    method_parts: list[str] = []
    is_group = False
    for token in tokens:
        if _is_condition_token(token):
            field, value = _split_condition(token)
            conditions.append((field, value))
            continue
        if token in GROUP_RULES:
            is_group = True
            continue
        if token in METHOD_WORDS:
            method_parts.append(token)
            continue
        source_parts.append(token)
    source = " ".join(source_parts).strip() or header
    method = _rule_method(" ".join(method_parts) or text)
    if is_group:
        method = "first"
    return RuleSpec(source=source, method=method, conditions=conditions, is_group=is_group)


def _is_condition_token(token: str) -> bool:
    return ("=" in token or "＝" in token) and not token.startswith("=")


def _split_condition(token: str) -> tuple[str, str]:
    if "＝" in token:
        field, value = token.split("＝", 1)
    else:
        field, value = token.split("=", 1)
    return _norm_rule(field), _norm_rule(value)


def _filter_data(data: pd.DataFrame, conditions: list[tuple[str, str]]) -> pd.DataFrame:
    if not conditions or data.empty:
        return data
    mask = pd.Series(True, index=data.index)
    for field, expected in conditions:
        source = _source_field(data, field)
        if source not in data.columns:
            mask &= False
            continue
        values = data[source].map(_text_value)
        expected_text = _business_level(expected) if source in {"牌级", "牌级别", "门店牌级"} else expected
        mask &= values == expected_text
    return data[mask]


def _align_grouped_series(series: pd.Series, group_fields: list[str], base_keys: pd.DataFrame) -> pd.Series:
    if series.empty:
        return pd.Series([0] * len(base_keys))
    value_frame = series.rename("_value").reset_index()
    merged = base_keys.merge(value_frame, on=group_fields, how="left")
    return merged["_value"].fillna(0).reset_index(drop=True)


def _is_formula_rule(rule: str) -> bool:
    text = _normalize_formula_text(rule)
    if not text:
        return False
    if re.fullmatch(r"=?[A-Z]+\d+([+\-*/][A-Z]+\d+)+", text, flags=re.IGNORECASE):
        return False
    return any(op in text for op in ["/", "+", "-", "*"]) and not text.startswith("http")


def _evaluate_formula(rule: str, output: pd.DataFrame, grouped, group_fields: list[str], data: pd.DataFrame, row_count: int) -> pd.Series:
    base_keys = output[group_fields].copy()
    return _evaluate_expression(_normalize_formula_text(rule).lstrip("="), lambda operand: _formula_operand(operand, output, grouped, group_fields, base_keys, data, row_count), row_count)


def _evaluate_row_formula(rule: str, output: pd.DataFrame, data: pd.DataFrame) -> pd.Series:
    return _evaluate_expression(_normalize_formula_text(rule).lstrip("="), lambda operand: _row_formula_operand(operand, output, data), len(data))


def _evaluate_expression(text: str, operand_getter, row_count: int) -> pd.Series:
    text = _strip_outer_parens(text)
    split = _split_top_level_operator(text, ["+", "-"])
    if not split:
        split = _split_top_level_operator(text, ["*", "/"])
    if not split:
        return operand_getter(text)
    left_name, operator, right_name = split
    left = _evaluate_expression(left_name, operand_getter, row_count)
    right = _evaluate_expression(right_name, operand_getter, row_count)
    if operator == "/":
        numerator = pd.to_numeric(left, errors="coerce").fillna(0)
        denominator = pd.to_numeric(right, errors="coerce").fillna(0)
        return (numerator / denominator.where(denominator != 0)).fillna(0)
    if operator == "*":
        return left * right
    if operator == "+":
        return left + right
    if operator == "-":
        return left - right
    return pd.Series([0] * row_count)


def _formula_operand(name: str, output: pd.DataFrame, grouped, group_fields: list[str], base_keys: pd.DataFrame, data: pd.DataFrame, row_count: int) -> pd.Series:
    field, method, conditions = _formula_field_and_method(name)
    if field in output.columns:
        return pd.to_numeric(output[field], errors="coerce").fillna(0).reset_index(drop=True)
    source = _source_field(data, field)
    if source not in data.columns:
        return pd.Series([0] * row_count)
    rule_spec = RuleSpec(source=source, method=_rule_method(method or "求和") or "sum", conditions=conditions)
    return _aggregate_template_column(grouped, group_fields, base_keys, data, source, rule_spec)


def _row_formula_operand(name: str, output: pd.DataFrame, data: pd.DataFrame) -> pd.Series:
    field, _method, conditions = _formula_field_and_method(name)
    if field in output.columns:
        return pd.to_numeric(output[field], errors="coerce").fillna(0).reset_index(drop=True)
    filtered = _filter_data(data, conditions)
    source = _source_field(filtered, field)
    if source not in filtered.columns:
        return pd.Series([0] * len(data))
    return pd.to_numeric(filtered[source].reindex(data.index), errors="coerce").fillna(0).reset_index(drop=True)


def _formula_field_and_method(name: str) -> tuple[str, str, list[tuple[str, str]]]:
    text = _normalize_formula_text(name)
    parsed = _parse_rule(text, text)
    return parsed.source, parsed.method, parsed.conditions


def _normalize_formula_text(value: object) -> str:
    text = _norm_rule(value)
    return text.replace("（", "(").replace("）", ")").replace("，", ",").replace("＝", "=")


def _strip_outer_parens(text: str) -> str:
    text = text.strip()
    while text.startswith("(") and text.endswith(")") and _outer_parens_wrap(text):
        text = text[1:-1].strip()
    return text


def _outer_parens_wrap(text: str) -> bool:
    depth = 0
    for idx, char in enumerate(text):
        if char == "(":
            depth += 1
        elif char == ")":
            depth -= 1
            if depth == 0 and idx != len(text) - 1:
                return False
    return depth == 0


def _split_top_level_operator(text: str, operators: list[str]) -> tuple[str, str, str] | None:
    depth = 0
    for idx in range(len(text) - 1, -1, -1):
        char = text[idx]
        if char == ")":
            depth += 1
        elif char == "(":
            depth -= 1
        elif depth == 0 and char in operators and idx > 0:
            return text[:idx].strip(), char, text[idx + 1 :].strip()
    return None


def _project_template_columns(data: pd.DataFrame, headers: list[str]) -> pd.DataFrame:
    projected = pd.DataFrame(index=data.index)
    for header in headers:
        source = _source_field(data, header)
        projected[header] = data[source] if source in data.columns else ""
    return projected.reset_index(drop=True)


def _source_field(data: pd.DataFrame, header: str, rule: str = "") -> str:
    raw_header = _norm_rule(header)
    if raw_header in data.columns:
        return raw_header
    header = _normalize_field_name(raw_header)
    if header in data.columns:
        return header
    if header in FIELD_ALIASES and FIELD_ALIASES[header] in data.columns:
        return FIELD_ALIASES[header]
    stripped = re.sub(r"(平均值|最高值|最低值|求和|合计|汇总|平均|均值|最大|最高|最小|最低|取整)$", "", header)
    if stripped in data.columns:
        return stripped
    if stripped in FIELD_ALIASES and FIELD_ALIASES[stripped] in data.columns:
        return FIELD_ALIASES[stripped]
    rule_field = _norm_rule(rule)
    if rule_field in data.columns:
        return rule_field
    return header


def _normalize_field_name(value: object) -> str:
    text = _norm_rule(value)
    return FIELD_ALIASES.get(text, text)


def _source_for_sheet(name: str, result: ProcessResult) -> pd.DataFrame:
    if name.startswith("美团"):
        return result.details_by_platform.get("meituan", pd.DataFrame())
    if name.startswith("抖音"):
        return result.details_by_platform.get("douyin", pd.DataFrame())
    if "合并" in name:
        return result.combined_detail
    return result.combined_detail


def _uses_platform_header(name: str, platform: str | None = None) -> bool:
    if name == "处理说明" or name.endswith("明细"):
        return False
    if platform in {"meituan", "douyin"}:
        return True
    return name.startswith("美团") or name.startswith("抖音")


def _single_platform(result: ProcessResult) -> str | None:
    platforms = [platform for platform, detail in result.details_by_platform.items() if not detail.empty]
    return platforms[0] if len(platforms) == 1 else None


def _platform_from_sheet(name: str) -> str | None:
    if name.startswith("美团"):
        return "meituan"
    if name.startswith("抖音"):
        return "douyin"
    return None


def _resolve_template(template_path: str | Path | None, default_path: Path) -> Path | None:
    if template_path:
        path = Path(template_path)
        return path if path.exists() else None
    if default_path.exists():
        return default_path
    legacy = ROOT / "配置表" / "输出报表模板.xlsx"
    if legacy.exists():
        return legacy
    return None


def _result_for_platform(result: ProcessResult, platform: str) -> ProcessResult:
    detail = result.details_by_platform.get(platform, pd.DataFrame())
    period_start, period_end = _period_range_for_detail(detail)
    return ProcessResult(
        {platform: detail} if not detail.empty else {},
        detail,
        pd.DataFrame(),
        pd.DataFrame(),
        pd.DataFrame(),
        period_start,
        period_end,
        result.summary_fields,
        result.aggregate_map,
    )


def _period_range_for_detail(detail: pd.DataFrame) -> tuple[str, str]:
    if detail.empty or "统计日期" not in detail.columns:
        return "", ""
    dates = pd.to_datetime(detail["统计日期"], errors="coerce").dropna()
    if dates.empty:
        return "", ""
    return dates.min().date().isoformat(), dates.max().date().isoformat()


def _looks_numeric(series: pd.Series) -> bool:
    if series.empty:
        return False
    numeric = pd.to_numeric(series, errors="coerce")
    return numeric.notna().mean() >= 0.8


def _write_platform_header_sheet(wb, name: str, df: pd.DataFrame, result: ProcessResult, platform: str | None) -> None:
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
    _clear_sheet(ws)
    end_col = max(len(df.columns), 1)
    header_fill = MEITUAN_YELLOW if platform == "meituan" else DOUYIN_BLACK
    title_color = "111827" if platform == "meituan" else "FFFFFF"

    for col in range(1, end_col + 1):
        ws.cell(1, col).fill = PatternFill("solid", fgColor=header_fill)
        ws.cell(2, col).fill = PatternFill("solid", fgColor="FFFFFF")

    ws.row_dimensions[1].height = 43
    ws.row_dimensions[2].height = 24
    _add_platform_logo(ws, platform)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = name
    ws["A1"].font = Font(name=DEFAULT_FONT_NAME, bold=True, color=title_color, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws["A2"] = _period_label(result)
    ws["A2"].font = Font(name=DEFAULT_FONT_NAME, color="374151", size=11)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    _write_sheet(wb, name, df, start_row=3, clear=False)


def _write_sheet(wb, name: str, df: pd.DataFrame, start_row: int = 1, clear: bool = True) -> None:
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
    if clear:
        _clear_sheet(ws)
    if df.empty:
        if len(df.columns) > 0:
            for col_idx, column in enumerate(df.columns, 1):
                ws.cell(start_row, col_idx, column)
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[start_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = f"A{start_row + 1}"
            last_col = get_column_letter(len(df.columns))
            ws.auto_filter.ref = f"A{start_row}:{last_col}{start_row}"
        else:
            ws.cell(start_row, 1, "暂无数据")
            ws.cell(start_row, 1).font = Font(bold=True, color="666666")
        return

    for col_idx, column in enumerate(df.columns, 1):
        ws.cell(start_row, col_idx, column)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[start_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx, _excel_value(value))

    ws.freeze_panes = f"A{start_row + 1}"
    last_col = get_column_letter(len(df.columns))
    last_row = start_row + len(df)
    ws.auto_filter.ref = f"A{start_row}:{last_col}{last_row}"
    for idx, column in enumerate(df.columns, 1):
        max_len = max([len(str(column))] + [len(str(value)) for value in df[column].head(200).fillna("")])
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 36)


def _add_platform_logo(ws, platform: str | None) -> None:
    logo_path = DOUYIN_LOGO if platform == "douyin" else MEITUAN_LOGO
    if not logo_path.exists():
        return
    try:
        logo = OpenpyxlImage(str(logo_path))
    except Exception:
        return
    target_height_px = 57  # 1.5 cm at roughly 96 dpi.
    if logo.height:
        ratio = target_height_px / logo.height
        logo.height = target_height_px
        logo.width = int(logo.width * ratio)
    logo.anchor = "A1"
    ws.add_image(logo)


def _period_label(result: ProcessResult) -> str:
    start = _date_mmdd(result.period_start)
    end = _date_mmdd(result.period_end)
    if start and end:
        return f"统计周期：{start}-{end}"
    return "统计周期：未识别"


def _date_mmdd(value: str) -> str:
    if not value:
        return ""
    try:
        date = pd.to_datetime(value, errors="coerce")
        if pd.isna(date):
            return ""
        return date.strftime("%m.%d")
    except Exception:
        return ""


def _apply_default_font(wb) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                current = copy(cell.font)
                current.name = DEFAULT_FONT_NAME
                cell.font = current


def _clear_sheet(ws) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def _metric_label(field: str, method: str) -> str:
    labels = {"sum": "求和", "mean": "平均", "count": "计数", "none": "不汇总"}
    return f"{field}({labels.get(method, method)})"


def _excel_value(value):
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    return value


def _number_column(df: pd.DataFrame, field: str) -> pd.Series:
    if field not in df.columns:
        return pd.Series(0, index=df.index, dtype="float64")
    return pd.to_numeric(df[field], errors="coerce").fillna(0)


def _text_value(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).replace("\u3000", " ").strip()


def _norm_rule(value: object) -> str:
    return " ".join(_text_value(value).split()).strip()


def _business_level(value: object) -> str:
    text = _text_value(value)
    return text if text in BUSINESS_LEVELS[:3] else "无等级"


def _first_non_empty(df: pd.DataFrame, columns: list[str]) -> pd.Series:
    result = pd.Series("", index=df.index, dtype="object")
    for column in columns:
        if column not in df.columns:
            continue
        values = df[column].map(_text_value)
        result = result.where(result.map(_text_value) != "", values)
    return result


def _order_existing(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    for column in columns:
        if column not in df.columns:
            df[column] = ""
    return df[columns]
