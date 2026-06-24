from __future__ import annotations

import hashlib
from datetime import datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel
from sqlalchemy import or_
from sqlalchemy.orm import Session

from ..config import settings
from ..db import get_db
from ..importers.assignment_keys import assignment_key
from ..importers.store_assignment_importer import import_store_assignments
from ..importers.store_matching import STORE_TYPE_LABELS, normalize_store_type, sync_stores_from_area_assignments
from ..models import AreaAssignment, MetricValue, Store


router = APIRouter()


class StoreAssignmentUpdate(BaseModel):
    store_type: str = "unknown"
    region: str
    owner: str
    note: str | None = None


@router.post("/upload")
def upload_store_config(file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict[str, object]:
    content = file.file.read()
    digest = hashlib.sha256(content).hexdigest()
    upload_dir = settings.upload_dir / "store-config"
    upload_dir.mkdir(parents=True, exist_ok=True)
    safe_name = Path(file.filename or "store_config.xlsx").name
    storage_path = upload_dir / f"{digest[:12]}_{safe_name}"
    storage_path.write_bytes(content)
    try:
        stats = import_store_assignments(db, storage_path)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return {"status": "imported", "sha256": digest, "stats": stats}


@router.get("/stores")
def list_stores(
    status: str | None = None,
    q: str | None = None,
    city: str | None = None,
    region: str | None = None,
    limit: int = 200,
    db: Session = Depends(get_db),
) -> dict[str, object]:
    query = _store_query(db, status=status, q=q, city=city, region=region)
    stores = query.order_by(Store.assignment_status.desc(), Store.city.asc(), Store.name.asc()).limit(max(1, min(limit, 500))).all()
    return {
        "items": [_store_payload(store) for store in stores],
        "store_type_labels": STORE_TYPE_LABELS,
    }


@router.get("/stores/export")
def export_stores(
    status: str | None = None,
    q: str | None = None,
    city: str | None = None,
    region: str | None = None,
    limit: int = 10000,
    db: Session = Depends(get_db),
) -> StreamingResponse:
    query = _store_query(db, status=status, q=q, city=city, region=region)
    stores = query.order_by(Store.assignment_status.desc(), Store.city.asc(), Store.name.asc()).limit(max(1, min(limit, 10000))).all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "区域配置"
    headers = [
        "平台",
        "匹配状态",
        "门店编码",
        "门店名称",
        "所在省份",
        "所在城市",
        "门店性质",
        "大区",
        "负责人",
        "匹配来源",
        "置信度",
        "说明",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="EAF1FF")
    header_font = Font(bold=True, color="0826B8")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    platform_map = _store_platform_map(db, stores)
    for store in stores:
        payload = _store_payload(store)
        sheet.append(
            [
                platform_map.get(str(payload["store_code"] or ""), ""),
                _status_label(str(payload["assignment_status"] or "")),
                payload["store_code"],
                payload["name"],
                payload["province"],
                payload["city"],
                STORE_TYPE_LABELS.get(str(payload["store_type"] or ""), str(payload["store_type"] or "")),
                payload["region"],
                payload["owner"],
                payload["assignment_source"],
                payload["assignment_confidence"],
                payload["assignment_note"],
            ]
        )

    widths = [14, 12, 22, 34, 14, 14, 12, 14, 14, 16, 10, 44]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"门店配置导出_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@router.post("/stores/{store_code}/assignment")
def update_store_assignment(store_code: str, payload: StoreAssignmentUpdate, db: Session = Depends(get_db)) -> dict[str, object]:
    store = db.query(Store).filter(Store.store_code == store_code).first()
    if not store:
        raise HTTPException(status_code=404, detail="门店不存在")
    region = payload.region.strip()
    owner = payload.owner.strip()
    if not region or not owner:
        raise HTTPException(status_code=422, detail="大区和负责人不能为空")
    store_type = normalize_store_type(payload.store_type)
    if store_type == "all":
        store_type = "unknown"

    store.store_type = store_type
    store.region = region
    store.owner = owner
    store.assignment_status = "confirmed"
    store.assignment_source = "manual_store"
    store.assignment_confidence = 100
    store.assignment_note = payload.note or "人工确认门店归属"

    if store.province and store.city and store.name:
        _upsert_store_area_assignment(db, store=store, store_type=store_type, region=region, owner=owner)

    db.commit()
    db.refresh(store)
    return {"status": "updated", "store": _store_payload(store)}


@router.post("/stores/reapply")
def reapply_store_assignments(db: Session = Depends(get_db)) -> dict[str, object]:
    updated = sync_stores_from_area_assignments(db)
    db.commit()
    return {"status": "updated", "updated_stores": updated}


def _upsert_store_area_assignment(db: Session, *, store: Store, store_type: str, region: str, owner: str) -> None:
    target_key = (*assignment_key(store.province, store.city, store.name), store_type)
    candidates = db.query(AreaAssignment).filter(AreaAssignment.city.in_(_city_candidates(store.city))).all()
    existing = next(
        (
            item
            for item in candidates
            if (*assignment_key(item.province, item.city, item.store_name), normalize_store_type(item.store_type)) == target_key
        ),
        None,
    )
    if existing:
        existing.region = region
        existing.owner = owner
        existing.enabled = True
        return
    province, city, store_name, normalized_type = target_key
    db.add(
        AreaAssignment(
            province=province,
            city=city,
            store_name=store_name,
            store_type=normalized_type,
            region=region,
            owner=owner,
            enabled=True,
        )
    )


def _city_candidates(city: str | None) -> list[str]:
    text = "" if city is None else city.strip()
    if not text:
        return [""]
    candidates = {text}
    if text.endswith("市"):
        candidates.add(text[:-1])
    else:
        candidates.add(f"{text}市")
    return list(candidates)


def _store_query(db: Session, *, status: str | None, q: str | None, city: str | None, region: str | None):
    query = db.query(Store)
    if status:
        query = query.filter(Store.assignment_status == status)
    if city:
        query = query.filter(Store.city.ilike(f"%{city.strip()}%"))
    if region:
        query = query.filter(Store.region == region.strip())
    if q:
        pattern = f"%{q.strip()}%"
        query = query.filter(or_(Store.name.ilike(pattern), Store.store_code.ilike(pattern), Store.owner.ilike(pattern)))
    return query


def _status_label(status: str) -> str:
    return {
        "confirmed": "已确认",
        "auto": "自动匹配",
        "review": "待确认",
        "unconfigured": "未配置",
    }.get(status, status)


def _store_platform_map(db: Session, stores: list[Store]) -> dict[str, str]:
    store_codes = [store.store_code for store in stores if store.store_code]
    if not store_codes:
        return {}
    rows = (
        db.query(MetricValue.store_code, MetricValue.platform_code)
        .filter(MetricValue.store_code.in_(store_codes))
        .distinct()
        .all()
    )
    platform_map: dict[str, set[str]] = {}
    for store_code, platform_code in rows:
        platform_map.setdefault(store_code, set()).add(_platform_label(platform_code))
    return {store_code: "、".join(sorted(platforms)) for store_code, platforms in platform_map.items()}


def _platform_label(platform_code: str | None) -> str:
    return {"meituan": "美团", "douyin": "抖音"}.get(platform_code or "", platform_code or "")


def _store_payload(store: Store) -> dict[str, object]:
    return {
        "store_code": store.store_code,
        "name": store.name,
        "province": store.province,
        "city": store.city,
        "region": store.region,
        "owner": store.owner,
        "store_type": store.store_type,
        "assignment_status": store.assignment_status,
        "assignment_source": store.assignment_source,
        "assignment_confidence": store.assignment_confidence,
        "assignment_note": store.assignment_note,
    }
