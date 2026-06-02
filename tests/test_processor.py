from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from tg_reporter.config import load_config
from tg_reporter.excel_report import write_excel_report, write_platform_excel_reports
from tg_reporter.history import archive_result
from tg_reporter.images import generate_briefing_images, generate_ranking_images
from tg_reporter.processor import ProcessResult, process_reports


def write_config(path: Path) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                ["美团", "日期", "统计日期", "日期", "是", "否", "", "是"],
                ["美团", "门店名称", "门店名", "文本", "是", "否", "", "否"],
                ["美团", "门店ID", "门店ID", "文本", "是", "否", "", "否"],
                ["美团", "订单数", "订单数", "数值", "是", "是", "求和", "否"],
                ["美团", "销售额", "销售额", "数值", "是", "是", "求和", "否"],
                ["美团", "好评数", "好评数", "数值", "是", "是", "求和", "否"],
                ["美团", "美团星级", "星级", "数值", "是", "是", "平均", "否"],
                ["美团", "门店名称", "门店计数", "文本", "否", "是", "计数", "否"],
                ["抖音", "统计日期", "统计日期", "日期", "是", "否", "", "是"],
                ["抖音", "门店", "门店名", "文本", "是", "否", "", "否"],
                ["抖音", "门店ID", "门店ID", "文本", "是", "否", "", "否"],
                ["抖音", "支付订单数", "订单数", "数值", "是", "是", "求和", "否"],
                ["抖音", "支付金额", "销售额", "数值", "是", "是", "求和", "否"],
                ["抖音", "好评", "好评数", "数值", "是", "是", "求和", "否"],
                ["抖音", "评分", "星级", "数值", "是", "是", "平均", "否"],
                ["抖音", "门店", "门店计数", "文本", "否", "是", "计数", "否"],
            ],
            columns=["平台", "原始字段", "标准字段", "字段类型", "是否输出", "是否汇总", "汇总方式", "是否日期字段"],
        ).to_excel(writer, sheet_name="字段映射", index=False)
        pd.DataFrame(
            [["", "示例门店A", "1001", "张三", "华东"], ["", "示例门店B", "1002", "李四", "华南"]],
            columns=["平台", "门店名", "门店ID", "门店负责人", "大区"],
        ).to_excel(writer, sheet_name="门店对照", index=False)
        pd.DataFrame([["广东省", "广州市", "张三", "华东"], ["湖北省", "武汉市", "李四", "华南"]], columns=["所在省份", "所在城市", "运营经理", "大区"]).to_excel(
            writer, sheet_name="区域负责人", index=False
        )
        pd.DataFrame([["华东", "王五"], ["华南", "赵六"]], columns=["大区", "大区负责人"]).to_excel(
            writer, sheet_name="大区对照", index=False
        )
        pd.DataFrame(
            [
                ["大区销售额排行榜", "合并", "大区", "销售额", "desc", 5, "横版", "是", "元"],
                ["门店好评排行榜", "合并", "门店名", "好评数", "desc", 5, "竖版", "是", ""],
            ],
            columns=["榜单名称", "数据范围", "统计维度", "指标字段", "排序", "TopN", "输出尺寸", "是否启用", "单位"],
        ).to_excel(writer, sheet_name="排行榜配置", index=False)
        pd.DataFrame(
            [
                [
                    "门店评级与评价综合简报",
                    "合并",
                    "横版",
                    "是",
                    "门店名",
                    "大区",
                    "星级",
                    "星级",
                    "星级",
                    "门店计数",
                    "好评数",
                    "好评数",
                    "差评数",
                    "订单数",
                    5,
                ]
            ],
            columns=[
                "简报名称",
                "数据范围",
                "输出尺寸",
                "是否启用",
                "门店字段",
                "分组字段",
                "评分字段",
                "星级字段",
                "对比星级字段",
                "等级字段",
                "评价数字段",
                "好评数字段",
                "差评数字段",
                "订单字段",
                "TopN",
            ],
        ).to_excel(writer, sheet_name="综合简报", index=False)


def test_end_to_end_single_and_dual_platform(tmp_path: Path) -> None:
    config_path = tmp_path / "config.xlsx"
    meituan_path = tmp_path / "meituan.xlsx"
    douyin_path = tmp_path / "douyin.xlsx"
    write_config(config_path)
    pd.DataFrame(
        [["2026-05-01", "示例门店A", "1001", 10, 1000, 8, 4.6], ["2026-05-01", "未知门店", "9999", 2, 50, 1, 3.6]],
        columns=["日期", "门店名称", "门店ID", "订单数", "销售额", "好评数", "美团星级"],
    ).to_excel(meituan_path, index=False)
    pd.DataFrame(
        [["2026-05-01", "示例门店B", "1002", 7, 700, 6, 4.9]],
        columns=["统计日期", "门店", "门店ID", "支付订单数", "支付金额", "好评", "评分"],
    ).to_excel(douyin_path, index=False)

    config = load_config(config_path)
    result = process_reports(config, {"meituan": meituan_path, "douyin": douyin_path})

    assert len(result.combined_detail) == 3
    assert result.period_start == "2026-05-01"
    assert result.period_end == "2026-05-01"
    assert result.combined_detail["销售额"].sum() == 1750
    assert result.combined_detail.loc[0, "门店负责人"] == "张三"
    assert result.combined_detail.loc[0, "城市"] == "广州市"
    assert result.combined_detail.loc[0, "大区"] == "华东"
    assert result.aggregate_map["销售额"] == "sum"
    assert result.aggregate_map["星级"] == "mean"
    assert result.aggregate_map["门店计数"] == "count"
    assert round(float(result.combined_summary["星级"].mean()), 2) == 4.37
    assert int(result.combined_summary["门店计数"].sum()) == 3
    assert (result.combined_detail["处理状态"] == "异常").sum() == 1

    excel_path = write_excel_report(result, tmp_path / "out.xlsx")
    db_path = archive_result(result, tmp_path / "history.sqlite")
    image_paths = generate_ranking_images(result, config.rankings, tmp_path / "images")
    briefing_paths = generate_briefing_images(result, config.briefings, tmp_path / "briefings")

    assert excel_path.exists()
    assert db_path.exists()
    assert len(image_paths) == 2
    assert all(path.exists() for path in image_paths)
    assert len(briefing_paths) == 1
    assert all(path.exists() for path in briefing_paths)


def test_single_platform_allowed(tmp_path: Path) -> None:
    config_path = tmp_path / "config.xlsx"
    meituan_path = tmp_path / "meituan.xlsx"
    write_config(config_path)
    pd.DataFrame(
        [["2026-05-02", "示例门店A", "1001", 3, 300, 2, 4.8]],
        columns=["日期", "门店名称", "门店ID", "订单数", "销售额", "好评数", "美团星级"],
    ).to_excel(meituan_path, index=False)

    result = process_reports(load_config(config_path), {"meituan": meituan_path, "douyin": None})

    assert list(result.details_by_platform) == ["meituan"]
    assert len(result.combined_detail) == 1
    assert result.combined_detail.iloc[0]["平台"] == "美团"


def test_output_template_rule_row_controls_meituan_summary(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "out.xlsx"
    wb = Workbook()
    wb.active.title = "合并明细"
    for sheet in ["美团明细", "抖音明细", "美团流量汇总", "美团经营汇总", "抖音汇总", "合并汇总", "处理说明"]:
        wb.create_sheet(sheet)
    ws = wb["美团流量汇总"]
    ws.append(["说明"])
    ws.append(["说明"])
    ws.append(["平台", "大区", "曝光人数(人)", "访问人数(人)", "曝光访问率", "门店数"])
    ws.append(["平台 分组", "大区 分组", "曝光人数(人) 求和", "访问人数(人) 求和", "访问人数(人) 求和/曝光人数(人) 求和", "门店名 去重计数"])
    ws = wb["美团经营汇总"]
    ws.append(["说明"])
    ws.append(["说明"])
    ws.append(["平台", "大区", "牌级", "门店数", "新增评价", "新增好评", "好评率"])
    ws.append(["平台 分组", "大区 分组", "牌级 分组", "门店名 去重计数", "新增评价 求和", "新增好评 求和", "新增好评 求和/新增评价 求和"])
    wb.save(template_path)

    detail = pd.DataFrame(
        [
            {
                "平台": "美团",
                "统计日期": "2026-05-01",
                "大区": "华东",
                "门店名": "店A",
                "曝光人数(人)": 100,
                "访问人数(人)": 50,
                "新增评价": 10,
                "新增好评": 8,
                "牌级别": "金牌",
            },
            {
                "平台": "美团",
                "统计日期": "2026-05-02",
                "大区": "华东",
                "门店名": "店B",
                "曝光人数(人)": 300,
                "访问人数(人)": 150,
                "新增评价": 5,
                "新增好评": 4,
                "牌级别": "钻石",
            },
        ]
    )
    result = ProcessResult(
        {"meituan": detail},
        detail,
        pd.DataFrame(),
        pd.DataFrame(),
        pd.DataFrame(),
        "2026-05-01",
        "2026-05-02",
        ["曝光人数(人)", "访问人数(人)", "新增评价", "新增好评"],
        {"曝光人数(人)": "sum", "访问人数(人)": "sum", "新增评价": "sum", "新增好评": "sum"},
    )

    write_excel_report(result, output_path, template_path)

    wb = load_workbook(output_path, data_only=True)
    traffic = wb["美团流量汇总"]
    assert [traffic.cell(3, col).value for col in range(1, 7)] == ["平台", "大区", "曝光人数(人)", "访问人数(人)", "曝光访问率", "门店数"]
    assert traffic.cell(4, 3).value == 400
    assert traffic.cell(4, 4).value == 200
    assert traffic.cell(4, 5).value == 0.5
    assert traffic.cell(4, 6).value == 2

    business_values = [wb["美团经营汇总"].cell(row, 3).value for row in range(4, 6)]
    assert business_values == ["金牌", "无等级"]


def test_platform_outputs_use_separate_periods_and_douyin_header(tmp_path: Path) -> None:
    meituan_template = tmp_path / "meituan_template.xlsx"
    douyin_template = tmp_path / "douyin_template.xlsx"
    for path, detail_sheet, summary_sheet in [
        (meituan_template, "美团明细", "美团大区汇总"),
        (douyin_template, "抖音明细", "华北大区"),
    ]:
        wb = Workbook()
        wb.active.title = detail_sheet
        ws = wb.create_sheet(summary_sheet)
        ws.append(["说明"])
        ws.append(["说明"])
        ws.append(["大区", "订单数"])
        ws.append(["大区 分组", "订单数 求和"])
        wb.create_sheet("处理说明")
        wb.save(path)

    meituan_detail = pd.DataFrame(
        [
            {"平台": "美团", "统计日期": "2026-05-01", "大区": "华东", "门店名": "店A", "订单数": 1},
            {"平台": "美团", "统计日期": "2026-05-28", "大区": "华东", "门店名": "店B", "订单数": 2},
        ]
    )
    douyin_detail = pd.DataFrame(
        [
            {"平台": "抖音", "统计日期": "2026-06-05", "大区": "华北", "门店名": "店C", "订单数": 3},
            {"平台": "抖音", "统计日期": "2026-06-20", "大区": "华北", "门店名": "店D", "订单数": 4},
        ]
    )
    combined = pd.concat([meituan_detail, douyin_detail], ignore_index=True)
    result = ProcessResult(
        {"meituan": meituan_detail, "douyin": douyin_detail},
        combined,
        pd.DataFrame(),
        pd.DataFrame(),
        pd.DataFrame(),
        "2026-05-01",
        "2026-06-20",
        ["订单数"],
        {"订单数": "sum"},
    )

    paths = write_platform_excel_reports(result, tmp_path, "TEST", meituan_template, douyin_template)
    assert [path.name for path in paths] == ["美团报表输出_TEST.xlsx", "抖音报表输出_TEST.xlsx"]

    meituan_wb = load_workbook(paths[0])
    douyin_wb = load_workbook(paths[1])
    assert meituan_wb["美团大区汇总"]["A1"].value == "美团大区汇总"
    assert meituan_wb["美团大区汇总"]["A2"].value == "统计周期：05.01-05.28"
    assert douyin_wb["华北大区"]["A1"].value == "华北大区"
    assert douyin_wb["华北大区"]["A1"].fill.fgColor.rgb == "00000000"
    assert douyin_wb["华北大区"]["A1"].font.color.rgb == "00FFFFFF"
    assert douyin_wb["华北大区"]["A2"].value == "统计周期：06.05-06.20"
