from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]


def main() -> None:
    config_path = ROOT / "配置表" / "报表工具配置模板.xlsx"
    template_path = ROOT / "配置表" / "美团输出报表模板.xlsx"
    config_path.parent.mkdir(parents=True, exist_ok=True)
    build_config(config_path)
    build_output_template(template_path)
    print(f"配置模板已生成：{config_path}")
    print(f"输出模板已生成：{template_path}")


def build_config(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "区域负责人"
    write_table(
        ws,
        ["所在省份", "所在城市", "运营经理", "大区"],
        [
            ["广东省", "广州市", "杨昊", "华东"],
            ["湖北省", "武汉市", "胡金伟", "华中"],
            ["上海市", "上海市", "肖炳宇", "华北"],
        ],
    )

    ws = wb.create_sheet("字段映射")
    write_table(
        ws,
        ["平台", "原始字段", "标准字段", "字段类型", "是否输出", "是否汇总", "汇总方式", "是否日期字段"],
        [
            ["美团", "日期", "统计日期", "日期", "是", "否", "", "是"],
            ["美团", "门店名称", "门店名", "文本", "是", "否", "", "否"],
            ["美团", "门店ID", "门店ID", "文本", "是", "否", "", "否"],
            ["美团", "订单数", "订单数", "数值", "是", "是", "求和", "否"],
            ["美团", "销售额", "销售额", "数值", "是", "是", "求和", "否"],
            ["美团", "好评数", "好评数", "数值", "是", "是", "求和", "否"],
            ["美团", "美团星级(星)", "美团星级", "数值", "是", "是", "平均", "否"],
            ["美团", "经营评分牌级", "牌级别", "文本", "是", "是", "计数", "否"],
            ["美团", "门店名称", "门店计数", "文本", "否", "是", "计数", "否"],
            ["抖音", "统计日期", "统计日期", "日期", "是", "否", "", "是"],
            ["抖音", "门店", "门店名", "文本", "是", "否", "", "否"],
            ["抖音", "门店ID", "门店ID", "文本", "是", "否", "", "否"],
            ["抖音", "支付订单数", "订单数", "数值", "是", "是", "求和", "否"],
            ["抖音", "支付金额", "销售额", "数值", "是", "是", "求和", "否"],
            ["抖音", "好评", "好评数", "数值", "是", "是", "求和", "否"],
        ],
    )

    ws = wb.create_sheet("门店对照")
    write_table(
        ws,
        ["平台", "门店名", "门店ID", "省份", "城市"],
        [
            ["", "示例门店A", "1001", "广东省", "广州市"],
            ["", "示例门店B", "1002", "湖北省", "武汉市"],
        ],
    )

    ws = wb.create_sheet("排行榜配置")
    write_table(
        ws,
        ["榜单名称", "数据范围", "统计维度", "指标字段", "排序", "TopN", "输出尺寸", "是否启用", "单位", "筛选字段", "筛选值"],
        [
            ["大区销售额排行榜", "合并", "大区", "销售额", "desc", 10, "全部", "是", "元", "", ""],
            ["门店好评排行榜", "合并", "门店名", "好评数", "desc", 10, "全部", "是", "", "", ""],
            ["金牌大区门店数", "合并", "大区", "牌级别", "desc", 10, "全部", "否", "家", "牌级别", "金牌"],
        ],
    )

    ws = wb.create_sheet("综合简报")
    write_table(
        ws,
        [
            "简报名称",
            "数据范围",
            "输出尺寸",
            "是否启用",
            "门店字段",
            "分组字段",
            "评分字段",
            "星级字段",
            "对比星级字段",
            "等级字段",
            "评价数字段",
            "好评数字段",
            "差评数字段",
            "订单字段",
            "TopN",
        ],
        [
            [
                "门店评级与评价综合简报",
                "美团",
                "汇报横版",
                "是",
                "门店名",
                "大区",
                "经营评分",
                "美团星级",
                "点评星级",
                "牌级别",
                "新增评价",
                "新增好评",
                "新增差评",
                "核销单量",
                12,
            ]
        ],
    )
    wb.save(path)


def build_output_template(path: Path) -> None:
    wb = Workbook()
    wb.active.title = "合并明细"
    for sheet in ["美团明细", "抖音明细", "美团流量汇总", "美团经营汇总", "抖音汇总", "合并汇总", "处理说明"]:
        wb.create_sheet(sheet)
    write_template_rules(
        wb["美团流量汇总"],
        ["平台", "大区", "门店负责人", "门店名称", "曝光人数(人)", "访问人数(人)", "曝光访问率", "下单人数(人)", "核销单量", "新客核销（人）", "新增评价", "新增好评", "好评率"],
        ["平台 分组", "大区 分组", "门店负责人 分组", "门店名 分组", "曝光人数(人) 求和", "访问人数(人) 求和", "访问人数(人) 求和/曝光人数(人) 求和", "下单人数(人) 求和", "核销单量 求和", "新客核销（人） 求和", "新增评价 求和", "新增好评 求和", "新增好评 求和/新增评价 求和"],
        "第 3 行是最终输出表头；第 4 行是取值/计算规则。规则里出现“分组/分类/维度/纬度”就作为分组字段；指标写“字段名 算法 条件”，比例写“字段A 求和/字段B 求和”。",
    )
    write_template_rules(
        wb["美团经营汇总"],
        ["平台", "大区", "大区负责人", "牌级", "门店数", "新增评价", "新增好评", "好评率", "经营评分平均值", "经营评分最高值"],
        ["平台 分组", "大区 分组", "大区负责人 分组", "牌级 分组", "门店名 去重计数", "新增评价 求和", "新增好评 求和", "新增好评 求和/新增评价 求和", "经营评分 平均", "经营评分 最高"],
        "经营汇总会把牌级中的金牌/银牌/铜牌保留，其他值归并为“无等级”；横向统计金牌数量可写“门店名 去重计数 牌级=金牌”。",
    )
    wb.save(path)


def write_table(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for idx, header in enumerate(headers, 1):
        width = max(len(str(header)) + 4, 14)
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_template_rules(ws, headers: list[str], rules: list[str], note: str) -> None:
    ws.cell(1, 1, f"{ws.title} - 参考配置示例（实际生成报表时会自动覆盖）")
    ws.cell(2, 1, "第 1-2 行页眉、logo、周期、字体由代码控制；第 3 行是真正数据表头，第 4 行是计算规则。")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    rule_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(3, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, rule in enumerate(rules, 1):
        cell = ws.cell(4, col_idx, rule)
        cell.fill = rule_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(7, 1, "配置项")
    ws.cell(7, 2, "参考说明")
    ws.cell(8, 1, "模板读法")
    ws.cell(8, 2, note)
    ws.cell(8, 2).alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A5"
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(str(header)) + 4, 12), 24)


if __name__ == "__main__":
    main()
